import os
import json
import subprocess
import logging
import asyncio
import functools
from concurrent.futures import ThreadPoolExecutor
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor
from datetime import datetime
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.dispatcher.handler import CancelHandler
from aiogram.dispatcher.middlewares import BaseMiddleware
from openpyxl import load_workbook

# --- Настройки ---
BOT_TOKEN = "7597757770:AAEySxgVkO5ei3uOKU-V5VbuXbxp7wje68Y" # Укажите ваш токен
ADMIN_IDS = [7616566890] # Укажите ID администраторов

# --- Настройка логгирования (в файл и в консоль) ---
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bot.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# --- Глобальные переменные и пути ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INVOICE_DIR = os.path.join(BASE_DIR, "invoices")
DRAFTS_DIR = os.path.join(BASE_DIR, "drafts")
FAVORITES_DIR = os.path.join(BASE_DIR, "favorites")
PAID_USERS_FILE = os.path.join(BASE_DIR, "paid_users.json")
MAINTENANCE_FILE = os.path.join(BASE_DIR, "maintenance.json")

# Создаем необходимые директории при запуске
os.makedirs(INVOICE_DIR, exist_ok=True)
os.makedirs(DRAFTS_DIR, exist_ok=True)
os.makedirs(FAVORITES_DIR, exist_ok=True)

# Пул потоков для выполнения блокирующих операций
executor_pool = ThreadPoolExecutor(max_workers=5)

# --- FSM States (Состояния) ---
class Form(StatesGroup):
    choose_action = State()
    new_device_name = State()
    new_device_sn = State()
    new_device_qty = State()
    old_device_name = State()
    old_device_sn = State()
    old_device_qty = State()
    shop_number = State()
    confirmation = State()
    check_shop_number = State()
    ai_assistant = State()
    search_shops = State()

class AdminStates(StatesGroup):
    mailing_message = State()
    mailing_confirmation = State()


# --- Вспомогательные функции ---

async def show_typing(obj, duration=1):
    """Асинхронно показывает действие 'печатает'."""
    try:
        chat_id = obj.chat.id if hasattr(obj, "chat") else obj.message.chat.id
        bot = obj.bot if hasattr(obj, "bot") else obj.message.bot
        await bot.send_chat_action(chat_id, ChatActions.TYPING)
        await asyncio.sleep(duration)
    except Exception as e:
        logging.warning(f"show_typing error: {e}")

async def show_processing(message: types.Message, text: str = "⏳ Подождите, идет обработка..."):
    """Показывает временное сообщение о процессе и возвращает его для последующего удаления."""
    # Для инлайн-кнопок используем message.message, для обычных - message
    real_message = message.message if isinstance(message, types.CallbackQuery) else message
    msg = await real_message.answer(text)
    return msg

async def run_in_thread(func, *args, **kwargs):
    """
    Запускает синхронную функцию в отдельном потоке с поддержкой
    позиционных (*args) и именованных (**kwargs) аргументов.
    """
    loop = asyncio.get_event_loop()
    p = functools.partial(func, *args, **kwargs)
    return await loop.run_in_executor(executor_pool, p)

async def send_long_message(message: types.Message, text: str, parse_mode: str = "Markdown"):
    """
    "Умно" отправляет длинное сообщение, разбивая его по строкам,
    чтобы не ломать Markdown-разметку.
    """
    if len(text) <= 4096:
        await message.answer(text, parse_mode=parse_mode)
        return

    lines = text.split('\n')
    chunk = ""
    for line in lines:
        if len(chunk) + len(line) + 1 > 4096:
            await message.answer(chunk, parse_mode=parse_mode)
            chunk = line + "\n"
        else:
            chunk += line + "\n"
    
    if chunk:
        await message.answer(chunk, parse_mode=parse_mode)

def escape_md(text: str) -> str:
    """
    Экранирует специальные символы для старого Markdown в Telegram.
    """
    text = str(text or "")
    escape_chars = r'_*`['
    return "".join('\\' + char if char in escape_chars else char for char in text)


def is_admin(user_id):
    """Проверяет, является ли пользователь администратором."""
    return user_id in ADMIN_IDS

def query_tinyllama(prompt: str) -> str:
    """Выполняет запрос к локальной модели TinyLlama через ollama."""
    try:
        process = subprocess.run(
            ['ollama', 'run', 'tinyllama'],
            input=prompt.encode('utf-8'),
            capture_output=True,
            timeout=45
        )
        return process.stdout.decode('utf-8').strip()
    except subprocess.TimeoutExpired:
        return "⏰ Время ожидания ответа от AI истекло. Попробуйте еще раз."
    except Exception as e:
        logging.error(f"Ошибка при обращении к TinyLlama: {e}")
        return f"❌ Ошибка при обращении к TinyLlama: {e}"

def convert_excel_to_pdf(excel_path: str, pdf_path: str):
    """Конвертирует Excel файл в PDF с использованием LibreOffice."""
    try:
        output_dir = os.path.dirname(pdf_path)
        soffice_cmd = "soffice"
        if os.name == "nt":
            soffice_cmd = "C:\\Program Files\\LibreOffice\\program\\soffice.exe"
            if not os.path.exists(soffice_cmd):
                soffice_cmd = "soffice"

        subprocess.run([
            soffice_cmd,
            "--headless",
            "--convert-to", "pdf",
            "--outdir", output_dir,
            excel_path
        ], check=True, timeout=60)
    except FileNotFoundError:
        logging.error("Команда 'soffice' не найдена. Убедитесь, что LibreOffice установлен и добавлен в PATH.")
        raise
    except subprocess.CalledProcessError as e:
        logging.error(f"Ошибка LibreOffice при конвертации: {e}")
        raise
    except subprocess.TimeoutExpired:
        logging.error("Таймаут при конвертации Excel в PDF.")
        raise

# --- Функции режима технического обслуживания ---
def set_maintenance_mode(status: bool):
    """Включает или выключает режим тех. обслуживания."""
    with open(MAINTENANCE_FILE, 'w') as f:
        json.dump({'maintenance_on': status}, f)

def is_maintenance_mode_on():
    """Проверяет, включен ли режим тех. обслуживания."""
    if not os.path.exists(MAINTENANCE_FILE):
        return False
    with open(MAINTENANCE_FILE, 'r') as f:
        try:
            data = json.load(f)
            return data.get('maintenance_on', False)
        except json.JSONDecodeError:
            return False

# --- Middleware для режима тех. обслуживания ---
class MaintenanceMiddleware(BaseMiddleware):
    async def on_pre_process_update(self, update: types.Update, data: dict):
        if not is_maintenance_mode_on():
            return

        user_id = None
        if update.message:
            user_id = update.message.from_user.id
        elif update.callback_query:
            user_id = update.callback_query.from_user.id

        if user_id and not is_admin(user_id):
            if update.message:
                await update.message.answer("🛠️ Бот находится на техническом обслуживании. Пожалуйста, попробуйте позже.")
            elif update.callback_query:
                await update.callback_query.answer("Бот на техническом обслуживании.", show_alert=True)
            raise CancelHandler()

# --- Работа с данными пользователей ---

def load_paid_users():
    """Загружает список ID платных пользователей из JSON файла."""
    if not os.path.exists(PAID_USERS_FILE):
        return []
    try:
        with open(PAID_USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        return []

def save_paid_user(user_id):
    """Сохраняет ID нового платного пользователя."""
    users = load_paid_users()
    if user_id not in users:
        users.append(user_id)
        with open(PAID_USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, indent=4)

def save_draft(user_id: int, data: dict):
    """Сохраняет черновик накладной для пользователя."""
    draft_file = os.path.join(DRAFTS_DIR, f"{user_id}_drafts.json")
    drafts = load_drafts(user_id)
    drafts.append({
        "timestamp": datetime.now().isoformat(),
        "data": data
    })
    with open(draft_file, "w", encoding="utf-8") as f:
        json.dump(drafts[-10:], f, indent=4)

def load_drafts(user_id: int):
    """Загружает черновики пользователя."""
    draft_file = os.path.join(DRAFTS_DIR, f"{user_id}_drafts.json")
    if os.path.exists(draft_file):
        try:
            with open(draft_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError:
            return []
    return []

# --- Система обучения AI ---

def improve_ai_response(user_question: str, ai_response: str, user_feedback: str = None):
    """Логирует взаимодействие с AI для последующего анализа и дообучения."""
    log_file = os.path.join(BASE_DIR, "ai_learning_log.json")
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "question": user_question,
        "response": ai_response,
        "feedback": user_feedback
    }
    try:
        logs = []
        if os.path.exists(log_file):
            with open(log_file, "r", encoding="utf-8") as f:
                try:
                    logs = json.load(f)
                except json.JSONDecodeError:
                    logs = []
        logs.append(log_entry)
        with open(log_file, "w", encoding="utf-8") as f:
            json.dump(logs[-1000:], f, indent=4)
    except Exception as e:
        logging.error(f"Ошибка записи лога AI: {e}")


# --- Инициализация бота ---
bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)


# --- Работа с Excel ---

def get_shop_info_by_number(number: str):
    """Ищет информацию о магазине по его номеру в Excel файле."""
    shop_file = os.path.join(BASE_DIR, "список магазинов .xlsx")
    if not os.path.exists(shop_file):
        return ("Код не найден", "Адрес не найден")
    try:
        wb = load_workbook(shop_file, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5 and row[3] and str(row[3]).strip() == number.strip():
                code = str(row[2]).strip() if row[2] else "Код не найден"
                address = str(row[4]).strip() if row[4] else "Адрес не найден"
                return (code, address)
        return ("Код не найден", "Адрес не найден")
    except Exception as e:
        logging.error(f"Ошибка при чтении файла магазинов: {e}")
        return ("Код не найден", "Адрес не найден")

def search_shops(query: str):
    """Выполняет 'умный' поиск магазинов по названию, адресу или коду."""
    shop_file = os.path.join(BASE_DIR, "список магазинов .xlsx")
    results = []
    if not os.path.exists(shop_file):
        return results
    try:
        wb = load_workbook(shop_file, read_only=True)
        ws = wb.active
        query_lower = query.lower()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                number = str(row[3]).strip() if row[3] else ""
                address = str(row[4]).strip() if row[4] else ""
                code = str(row[2]).strip() if row[2] else ""
                if (query_lower in number.lower() or
                    query_lower in address.lower() or
                    query_lower in code.lower()):
                    results.append({"number": number, "address": address, "code": code})
                    if len(results) >= 10:
                        break
        return results
    except Exception as e:
        logging.error(f"Ошибка при поиске магазинов: {e}")
        return []

def fill_template_optimized(filepath: str, devices: list, from_whom: str, to_whom: str, code_cell: str, code_value: str, static_cell: str = "BC9", static_value: str = "Smt9") -> str:
    """Оптимизированно заполняет шаблон Excel данными и сохраняет результат."""
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        ws["G12"] = from_whom
        ws["CH12"] = to_whom
        ws["AH9"] = datetime.today().strftime("%d.%m.%Y")
        ws[static_cell] = static_value
        ws[code_cell] = code_value

        start_row = 17
        for i, device in enumerate(devices):
            row = start_row + i
            ws[f"Y{row}"] = device["name"]
            ws[f"DW{row}"] = device["sn"]
            ws[f"BN{row}"] = device["qty"]

        filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(filepath)}"
        path = os.path.join(INVOICE_DIR, filename)
        wb.save(path)
        return path

    except Exception as e:
        logging.error(f"Критическая ошибка при заполнении шаблона Excel: {e}")
        raise

# --- Клавиатуры ---
def main_keyboard(user_id=None):
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(KeyboardButton("🏗️ Создать накладную"), KeyboardButton("📋 Список магазинов"))
    kb.add(KeyboardButton("📍 Проверить адрес"), KeyboardButton("🔍 Поиск магазина"))
    kb.add(KeyboardButton("📄 Шаблоны"), KeyboardButton("📦 Мои накладные"))
    kb.add(KeyboardButton("🤖 AI-помощник"), KeyboardButton("⚙️ Настройки"))
    kb.add(KeyboardButton("ℹ️ О боте"), KeyboardButton("🆘 Поддержка"))
    if user_id and is_admin(user_id):
        kb.add(KeyboardButton("👨‍💻 Админ-панель"))
    return kb

def admin_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(KeyboardButton("📢 Рассылка"), KeyboardButton("📊 Логи AI"))
    kb.add(KeyboardButton("👥 Статистика"), KeyboardButton("💰 Платежи"))
    kb.add(KeyboardButton("🔄 Перезагрузка"), KeyboardButton("🔧 Вкл/Выкл тех. режим"))
    kb.add(KeyboardButton("⬅️ Назад в меню"))
    return kb

def get_invoice_keyboard():
    kb = InlineKeyboardMarkup(row_width=2)
    kb.row(InlineKeyboardButton("➕ Новое устройство", callback_data="add_new"), InlineKeyboardButton("➖ Возврат", callback_data="add_old"))
    kb.row(InlineKeyboardButton("🏪 Магазин", callback_data="set_shop"), InlineKeyboardButton("📋 Просмотр", callback_data="view_data"))
    kb.row(InlineKeyboardButton("✏️ Редактировать", callback_data="edit_data"), InlineKeyboardButton("🗑️ Очистить", callback_data="clear_all"))
    kb.row(InlineKeyboardButton("💾 Сохранить черновик", callback_data="save_draft"), InlineKeyboardButton("📂 Загрузить черновик", callback_data="load_draft"))
    kb.row(InlineKeyboardButton("✅ Готово", callback_data="finish"), InlineKeyboardButton("❌ Отмена", callback_data="cancel"))
    return kb

def get_templates_keyboard():
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(InlineKeyboardButton("📄 М-15 установка", callback_data="template_m15_install"), InlineKeyboardButton("📄 М-15 возврат", callback_data="template_m15_return"))
    kb.add(InlineKeyboardButton("📄 АВР X5", callback_data="template_avr"), InlineKeyboardButton("📄 Акт приема", callback_data="template_acceptance"))
    return kb

def get_confirmation_keyboard():
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(InlineKeyboardButton("✅ Да, создать", callback_data="confirm_yes"), InlineKeyboardButton("✏️ Редактировать", callback_data="confirm_edit"))
    kb.add(InlineKeyboardButton("💾 Сохранить черновик", callback_data="save_draft_confirm"), InlineKeyboardButton("❌ Отменить", callback_data="confirm_no"))
    return kb

def get_ai_suggestions():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.row("❓ Как создать накладную?")
    kb.row("🏪 Как найти магазин?")
    kb.row("📄 Какие шаблоны есть?")
    kb.row("🚪 Выход из помощника")
    return kb

def get_ai_feedback_keyboard():
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(InlineKeyboardButton("👍 Полезно", callback_data="ai_feedback_good"), InlineKeyboardButton("👎 Не помогло", callback_data="ai_feedback_bad"))
    return kb

# --- AI-помощник ---
async def ai_assistant(message: types.Message, state: FSMContext = None):
    process_msg = await show_processing(message, "⏳ AI-помощник анализирует запрос...")
    try:
        user_question = message.text
        context = f"""<|system|>
Ты - AI-помощник в телеграм-боте для создания накладных М-15. Отвечай на русском языке кратко и понятно.

Функции бота:
- Создание накладных М-15 (установка и возврат оборудования)
- Поиск магазинов по базе
- Работа с шаблонами документов
- Проверка адресов магазинов

Текущее время: {datetime.now().strftime("%d.%m.%Y %H:%M")}
Пользователь: {message.from_user.full_name}

Вопрос: {user_question}

Отвечай полезно и по делу, используй эмодзи. Если вопрос не по теме бота, вежливо откажись отвечать.</s>
<|user|>
{user_question}</s>
<|assistant|>
"""
        response = await run_in_thread(query_tinyllama, context)
        if len(response) > 1000:
            response = response[:1000] + "..."

        formatted_response = f"🤖 *AI-помощник (TinyLlama):*\n\n{response}"
        await message.answer(formatted_response, parse_mode="Markdown", reply_markup=get_ai_feedback_keyboard())
        improve_ai_response(user_question, response)

    except Exception as e:
        logging.error(f"Ошибка AI-помощника: {e}")
        await message.answer("❌ Извините, произошла ошибка. Попробуйте позже или обратитесь в поддержку.")
    finally:
        await process_msg.delete()

# --- Основные обработчики ---

@dp.message_handler(commands="start")
async def cmd_start(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer(f"🤖 Добро пожаловать, {message.from_user.full_name}! Выберите действие:", reply_markup=main_keyboard(message.from_user.id))

@dp.message_handler(lambda m: m.text == "🏗️ Создать накладную", state="*")
async def create_invoice(message: types.Message, state: FSMContext):
    await state.finish()
    await state.update_data(new_devices=[], old_devices=[])
    text = (
        "🏗️ *Создание накладной М-15*\n\n"
        "Используйте кнопки ниже для добавления данных. Когда все будет готово, нажмите '✅ Готово'."
    )
    await message.answer(text, parse_mode="Markdown", reply_markup=get_invoice_keyboard())
    await Form.choose_action.set()

# --- Обработчики процесса создания накладной ---

@dp.callback_query_handler(state="*")
async def handle_inline_buttons(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    data = await state.get_data()
    action = callback_query.data

    if action == "add_new":
        await callback_query.message.answer("📦 Введите название нового устройства:")
        await Form.new_device_name.set()
    elif action == "add_old":
        await callback_query.message.answer("♻️ Введите название устройства для возврата:")
        await Form.old_device_name.set()
    elif action == "set_shop":
        await callback_query.message.answer("🏪 Введите номер магазина:")
        await Form.shop_number.set()
    
    # <<< ИЗМЕНЕНИЕ: ДОБАВЛЕН ИНДИКАТОР ПРОЦЕССА >>>
    elif action == "view_data":
        process_msg = await show_processing(callback_query, "⏳ Обновляю данные...")
        try:
            await show_current_data(callback_query.message, state)
        finally:
            await process_msg.delete()

    elif action == "finish":
        if not data.get("shop_number"):
            await callback_query.message.answer("❌ Сначала укажите номер магазина!")
            return
        if not data.get("new_devices") and not data.get("old_devices"):
            await callback_query.message.answer("❌ Добавьте хотя бы одно устройство!")
            return
        
        process_msg = await show_processing(callback_query, "⏳ Формирую итоговую сводку...")
        try:
            await show_summary(callback_query.message, state)
            await Form.confirmation.set()
        finally:
            await process_msg.delete()

    # <<< ИЗМЕНЕНИЕ: ДОБАВЛЕН ИНДИКАТОР ПРОЦЕССА >>>
    elif action == "clear_all":
        process_msg = await show_processing(callback_query, "⏳ Очищаю данные...")
        try:
            await state.update_data(new_devices=[], old_devices=[], shop_number=None)
            await callback_query.message.answer("🗑️ Все данные очищены!")
            await show_current_data(callback_query.message, state)
        finally:
            await process_msg.delete()

    elif action.startswith("template_"):
        await handle_template_callback(callback_query)
    elif action == "confirm_yes":
        await generate_docs_async(callback_query.message, state)
    elif action in ["confirm_no", "confirm_edit", "cancel"]:
        await callback_query.message.answer("Выберите действие:", reply_markup=get_invoice_keyboard())
        await Form.choose_action.set()
    elif action.startswith('ai_feedback_'):
        feedback_type = action.split('_')[-1]
        question = "Неизвестно"
        if callback_query.message.reply_to_message and callback_query.message.reply_to_message.from_user.id == callback_query.from_user.id:
            question = callback_query.message.reply_to_message.text
        
        ai_response = callback_query.message.text

        improve_ai_response(question, ai_response, feedback_type)
        feedback_text = "✅ Спасибо за обратную связь! Рад, что помог!" if feedback_type == "good" else "🙏 Спасибо! Постараюсь улучшить ответы."
        await callback_query.message.edit_text(callback_query.message.text, parse_mode="Markdown")
        await callback_query.message.answer(feedback_text)

async def handle_template_callback(callback_query: types.CallbackQuery):
    template_type = callback_query.data.split('_', 1)[1]
    process_msg = await show_processing(callback_query, "⏳ Ищу шаблон...")
    
    templates = {
        "m15_install": ("М-15 установка.xlsx", "📄 Шаблон М-15 (Установка)"),
        "m15_return": ("М-15 Возврат.xlsx", "📄 Шаблон М-15 (Возврат)"),
        "avr": ("АВР Х5 БС пдф.pdf", "📄 Акт выполненных работ (АВР X5)"),
    }
    
    if template_type in templates:
        file_name, caption = templates[template_type]
        file_path = os.path.join(BASE_DIR, file_name)
        if os.path.exists(file_path):
            await callback_query.message.answer_document(types.InputFile(file_path), caption=caption)
        else:
            await callback_query.message.answer(f"❌ Файл {file_name} не найден.")
    else:
        await callback_query.message.answer("⚠️ Неизвестный шаблон.")
        
    await process_msg.delete()


async def show_current_data(message: types.Message, state: FSMContext):
    data = await state.get_data()
    msg = "📊 *Текущие данные:*\n\n"
    if data.get("new_devices"):
        msg += "🆕 *Новые устройства:*\n"
        for i, dev in enumerate(data["new_devices"], 1):
            msg += f"{i}. {dev['name']} | SN: {dev['sn']} | Кол-во: {dev['qty']}\n"
    else:
        msg += "🆕 *Новые устройства:* нет\n"
    msg += "\n"

    if data.get("old_devices"):
        msg += "♻️ *Возврат:*\n"
        for i, dev in enumerate(data["old_devices"], 1):
            msg += f"{i}. {dev['name']} | SN: {dev['sn']} | Кол-во: {dev['qty']}\n"
    else:
        msg += "♻️ *Возврат:* нет\n"
    msg += "\n"
    
    shop_number = data.get("shop_number")
    if shop_number:
        code, address = get_shop_info_by_number(shop_number)
        msg += f"🏪 *Магазин:* №{shop_number} (Код: {code})\n📍 *Адрес:* {address}"
    else:
        msg += "🏪 *Магазин:* не указан"
    
    await message.answer(msg, parse_mode="Markdown", reply_markup=get_invoice_keyboard())

async def show_summary(message: types.Message, state: FSMContext):
    data = await state.get_data()
    summary = "📋 *Итоговые данные для накладной:*\n\n"
    if data.get("new_devices"):
        summary += "🆕 *Новые устройства:*\n"
        for i, dev in enumerate(data["new_devices"], 1):
            summary += f"{i}. {dev['name']} | SN: {dev['sn']} | Кол-во: {dev['qty']}\n"
        summary += "\n"
    if data.get("old_devices"):
        summary += "♻️ *Возврат:*\n"
        for i, dev in enumerate(data["old_devices"], 1):
            summary += f"{i}. {dev['name']} | SN: {dev['sn']} | Кол-во: {dev['qty']}\n"
        summary += "\n"
    shop_number = data.get("shop_number")
    code, address = get_shop_info_by_number(shop_number)
    summary += f"🏪 *Магазин:* №{shop_number}\n📍 *Адрес:* {address}\n\n"
    summary += "✅ Все верно? Создаем накладные?"
    await message.answer(summary, parse_mode="Markdown", reply_markup=get_confirmation_keyboard())


# --- Обработчики ввода данных для накладной ---

@dp.message_handler(state=Form.new_device_name)
async def input_new_name(message: types.Message, state: FSMContext):
    await state.update_data(current_new_name=message.text)
    await message.answer("🔢 Введите серийный номер нового устройства:")
    await Form.new_device_sn.set()

@dp.message_handler(state=Form.new_device_sn)
async def input_new_sn(message: types.Message, state: FSMContext):
    await state.update_data(current_new_sn=message.text)
    await message.answer("📦 Введите количество (только цифры):")
    await Form.new_device_qty.set()

@dp.message_handler(state=Form.new_device_qty)
async def save_new_device(message: types.Message, state: FSMContext):
    if not message.text.strip().isdigit() or int(message.text.strip()) <= 0:
        await message.answer("❌ Ошибка. Введите корректное количество (целое положительное число).")
        return
    
    process_msg = await show_processing(message, "⏳ Добавляю устройство...")
    data = await state.get_data()
    new_devices = data.get("new_devices", [])
    new_devices.append({
        "name": data["current_new_name"],
        "sn": data["current_new_sn"],
        "qty": int(message.text.strip())
    })
    await state.update_data(new_devices=new_devices)
    await message.answer(f"✅ Добавлено новое устройство: {data['current_new_name']}")
    
    await process_msg.delete()
    await show_current_data(message, state)
    await Form.choose_action.set()

@dp.message_handler(state=Form.old_device_name)
async def input_old_name(message: types.Message, state: FSMContext):
    await state.update_data(current_old_name=message.text)
    await message.answer("🔢 Введите серийный номер устройства для возврата:")
    await Form.old_device_sn.set()

@dp.message_handler(state=Form.old_device_sn)
async def input_old_sn(message: types.Message, state: FSMContext):
    await state.update_data(current_old_sn=message.text)
    await message.answer("📦 Введите количество для возврата (только цифры):")
    await Form.old_device_qty.set()

@dp.message_handler(state=Form.old_device_qty)
async def save_old_device(message: types.Message, state: FSMContext):
    if not message.text.strip().isdigit() or int(message.text.strip()) <= 0:
        await message.answer("❌ Ошибка. Введите корректное количество (целое положительное число).")
        return

    process_msg = await show_processing(message, "⏳ Добавляю возврат...")
    data = await state.get_data()
    old_devices = data.get("old_devices", [])
    old_devices.append({
        "name": data["current_old_name"],
        "sn": data["current_old_sn"],
        "qty": int(message.text.strip())
    })
    await state.update_data(old_devices=old_devices)
    await message.answer(f"✅ Добавлен возврат: {data['current_old_name']}")
    
    await process_msg.delete()
    await show_current_data(message, state)
    await Form.choose_action.set()

@dp.message_handler(state=Form.shop_number)
async def input_shop_number(message: types.Message, state: FSMContext):
    process_msg = await show_processing(message, "⏳ Ищу магазин в базе...")
    shop_number = message.text.strip()
    code, address = get_shop_info_by_number(shop_number)
    await process_msg.delete()

    if code == "Код не найден":
        await message.answer("❌ Магазин не найден. Проверьте номер и попробуйте снова:")
        return
    
    await state.update_data(shop_number=shop_number)
    await message.answer(f"✅ Магазин №{shop_number} сохранен!\n📍 Адрес: {address}")
    await show_current_data(message, state)
    await Form.choose_action.set()

# --- Генерация документов ---
async def generate_docs_async(message: types.Message, state: FSMContext):
    process_msg = await show_processing(message, "⏳ Начинаю генерацию документов...")
    
    try:
        data = await state.get_data()
        shop_number = data['shop_number']
        shop_code, shop_address = get_shop_info_by_number(shop_number)
        
        if data.get("new_devices"):
            await message.answer("📄 Создаю накладную на установку...")
            excel_path = await run_in_thread(
                fill_template_optimized,
                os.path.join(BASE_DIR, "М-15 установка.xlsx"),
                data["new_devices"],
                shop_address,
                "ООО ПКФ Бизнес Содействие",
                "CG9", shop_code
            )
            pdf_path = excel_path.replace(".xlsx", ".pdf")
            await run_in_thread(convert_excel_to_pdf, excel_path, pdf_path)
            await message.answer_document(types.InputFile(pdf_path), caption="📄 Накладная М-15 (Установка)")
            await run_in_thread(os.remove, excel_path)
            await run_in_thread(os.remove, pdf_path)

        if data.get("old_devices"):
            await message.answer("📄 Создаю накладную на возврат...")
            excel_path = await run_in_thread(
                fill_template_optimized,
                os.path.join(BASE_DIR, "М-15 Возврат.xlsx"),
                data["old_devices"],
                "ООО ПКФ Бизнес Содействие",
                shop_address,
                code_cell="BC9",
                code_value=shop_code,
                static_cell="CG9",
                static_value="Smt9"
            )
            pdf_path = excel_path.replace(".xlsx", ".pdf")
            await run_in_thread(convert_excel_to_pdf, excel_path, pdf_path)
            await message.answer_document(types.InputFile(pdf_path), caption="📄 Накладная М-15 (Возврат)")
            await run_in_thread(os.remove, excel_path)
            await run_in_thread(os.remove, pdf_path)

        await message.answer("✅ Накладные успешно созданы! Что дальше?", reply_markup=main_keyboard(message.from_user.id))

    except Exception as e:
        logging.exception("Критическая ошибка при создании накладных")
        await message.answer(f"❌ Произошла ошибка при создании накладных. Обратитесь в поддержку.\n\n`{e}`")
    finally:
        await process_msg.delete()
        await state.finish()

# --- Обработчики глобальных команд из меню ---

@dp.message_handler(lambda m: m.text == "📋 Список магазинов", state="*")
async def handle_shop_list(message: types.Message, state: FSMContext):
    await send_shop_list(message)

@dp.message_handler(lambda m: m.text == "📍 Проверить адрес", state="*")
async def handle_check_address(message: types.Message, state: FSMContext):
    await ask_shop_for_check(message)

@dp.message_handler(lambda m: m.text == "📄 Шаблоны", state="*")
async def handle_templates(message: types.Message, state: FSMContext):
    await show_templates(message)

@dp.message_handler(lambda m: m.text == "📦 Мои накладные", state="*")
async def handle_my_invoices(message: types.Message, state: FSMContext):
    await show_my_invoices(message)

@dp.message_handler(lambda m: m.text == "⚙️ Настройки", state="*")
async def handle_settings(message: types.Message, state: FSMContext):
    await show_settings(message)

@dp.message_handler(lambda m: m.text == "ℹ️ О боте", state="*")
async def handle_about(message: types.Message, state: FSMContext):
    await bot_info(message)

@dp.message_handler(lambda m: m.text == "🆘 Поддержка", state="*")
async def handle_support(message: types.Message, state: FSMContext):
    await support_info(message)


# --- Реализация команд из меню ---

async def send_shop_list(message: types.Message):
    """Отправляет пользователю список магазинов."""
    process_msg = await show_processing(message, "⏳ Загружаю список магазинов...")
    shop_file = os.path.join(BASE_DIR, "список магазинов .xlsx")
    if not os.path.exists(shop_file):
        await process_msg.delete()
        await message.answer("❌ Файл со списком магазинов не найден.")
        return

    try:
        wb = load_workbook(shop_file, read_only=True, data_only=True)
        ws = wb.active
        result = "*📋 Список магазинов:*\n\n"
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                number = escape_md(str(row[3]).strip() if row[3] else "—")
                address = escape_md(str(row[4]).strip() if row[4] else "—")
                result += f"🏪 *{number}* — {address}\n"
        
        await send_long_message(message, result, parse_mode="Markdown")

    except Exception as e:
        logging.error(f"Ошибка при чтении файла магазинов: {e}")
        await message.answer(f"❌ Ошибка при чтении файла: {e}")
    finally:
        await process_msg.delete()

async def ask_shop_for_check(message: types.Message):
    """Спрашивает номер магазина для проверки адреса."""
    await Form.check_shop_number.set()
    await message.answer("🔍 Введите номер магазина для проверки адреса:")

@dp.message_handler(state=Form.check_shop_number)
async def show_shop_address(message: types.Message, state: FSMContext):
    """Показывает адрес и код магазина по номеру."""
    process_msg = await show_processing(message, "⏳ Ищу магазин...")
    shop_number = message.text.strip()
    shop_code, shop_address = get_shop_info_by_number(shop_number)
    await process_msg.delete()
    if shop_code == "Код не найден":
        await message.answer("❌ Магазин не найден. Проверьте номер.")
    else:
        await message.answer(f"🏪 Магазин №{shop_number}\n📍 Адрес: {shop_address}\n🔢 Код: {shop_code}")
    await state.finish()

async def show_templates(message: types.Message):
    """Показывает клавиатуру с шаблонами."""
    await message.answer("📄 Выберите шаблон для скачивания:", reply_markup=get_templates_keyboard())

async def show_my_invoices(message: types.Message):
    """Показывает последние 5 черновиков."""
    process_msg = await show_processing(message, "⏳ Загружаю историю черновиков...")
    drafts = load_drafts(message.from_user.id)
    if drafts:
        response = "📦 *Ваши последние 5 черновиков:*\n\n"
        for i, draft in enumerate(drafts[-5:], 1):
            date = datetime.fromisoformat(draft['timestamp']).strftime("%d.%m.%Y %H:%M")
            response += f"{i}. 📅 {date}\n"
        response += "\n💡 Вы можете загрузить их в меню создания накладной."
        await message.answer(response, parse_mode="Markdown")
    else:
        await message.answer("📋 У вас пока нет сохраненных черновиков.")
    await process_msg.delete()

async def show_settings(message: types.Message):
    """Показывает информацию о настройках (заглушка)."""
    settings_text = "⚙️ *Настройки бота*\n\nНа данный момент все настройки управляются автоматически. В будущем здесь появятся опции для персонализации."
    await message.answer(settings_text, parse_mode="Markdown")

async def bot_info(message: types.Message):
    """Показывает информацию о боте."""
    bot_info_text = f"""
🤖 *Информация о боте*

*Название:* Бот для создания накладных М-15
*Версия:* 2.7 (Финальная)

*Основные функции:*
• 🏗️ Создание накладных М-15 (установка и возврат)
• 📄 Работа с шаблонами документов
• 🏪 Поиск и проверка магазинов в базе
• 🤖 AI-помощник для консультаций

*Технические особенности:*
• Автоматическая конвертация Excel → PDF
• Интеллектуальный поиск магазинов
• Поддержка многопоточности для тяжелых операций

*Системные требования:*
• LibreOffice для конвертации документов
• Ollama с моделью *TinyLlama* для AI-помощника

*Статус системы:* 🟢 Работает нормально
*Последнее обновление:* {datetime.now().strftime("%d.%m.%Y")}

Для технической поддержки используйте кнопку '🆘 Поддержка'.
    """
    await message.answer(bot_info_text, parse_mode="Markdown")

async def support_info(message: types.Message):
    """Показывает контакты поддержки."""
    support_text = (
        "🆘 *Техническая поддержка:*\n\n"
        "Если у вас возникли проблемы или есть предложения, свяжитесь с:\n\n"
        "👨‍💻 *Разработчик:* Рафаэль Абдуллаев\n"
        "🔗 @Rafael005t\n\n"
        "👨‍💼 *Администратор:* Намиг Гаджибеков\n"
        "🔗 @namig85\n\n"
    )
    await message.answer(support_text, parse_mode="Markdown")


# --- Поиск и AI ---

@dp.message_handler(lambda m: m.text == "🔍 Поиск магазина", state="*")
async def handle_search_shop(message: types.Message):
    """Начинает процесс поиска магазина."""
    await message.answer("🔍 Введите название, адрес или код магазина для поиска:")
    await Form.search_shops.set()

@dp.message_handler(state=Form.search_shops)
async def handle_search_query(message: types.Message, state: FSMContext):
    """Обрабатывает поисковый запрос и выдает результаты."""
    process_msg = await show_processing(message, "⏳ Ищу магазины...")
    query = message.text.strip()
    if len(query) < 2:
        await process_msg.delete()
        await message.answer("❌ Слишком короткий запрос. Введите хотя бы 2 символа.")
        return
    
    results = await run_in_thread(search_shops, query)
    await process_msg.delete()
    
    if not results:
        await message.answer("❌ Магазины не найдены. Попробуйте другой запрос.")
        return
    
    response = "🔍 *Результаты поиска:*\n\n"
    for shop in results:
        response += f"🏪 №{shop['number']}\n   📍 {shop['address']}\n   🔢 Код: {shop['code']}\n\n"
    await message.answer(response, parse_mode="Markdown")
    await state.finish()

@dp.message_handler(lambda m: m.text == "🤖 AI-помощник", state="*")
async def handle_ai_assistant_entry(message: types.Message, state: FSMContext):
    """Вход в режим AI-помощника."""
    await state.finish()
    welcome_text = """
🤖 *Добро пожаловать в AI-помощник!*

Я могу помочь вам с вопросами по работе бота. Просто задайте ваш вопрос, и я постараюсь помочь!

Или воспользуйтесь быстрыми командами ниже.
    """
    await message.answer(welcome_text, parse_mode="Markdown", reply_markup=get_ai_suggestions())
    await Form.ai_assistant.set()

@dp.message_handler(state=Form.ai_assistant)
async def handle_ai_question(message: types.Message, state: FSMContext):
    """Обрабатывает вопрос к AI или выход из режима."""
    if message.text == "🚪 Выход из помощника":
        await state.finish()
        await message.answer("👋 Возвращаюсь в главное меню!", reply_markup=main_keyboard(message.from_user.id))
        return
    await ai_assistant(message, state)
    

# --- Админ-панель ---

@dp.message_handler(lambda m: m.text == "👨‍💻 Админ-панель", state="*")
async def admin_panel(message: types.Message):
    if not is_admin(message.from_user.id): return
    await message.answer("👨‍💻 Панель администратора:", reply_markup=admin_keyboard())

@dp.message_handler(lambda m: m.text == "⬅️ Назад в меню", state="*")
async def back_to_main_menu(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await state.finish()
    await message.answer("Главное меню:", reply_markup=main_keyboard(message.from_user.id))

@dp.message_handler(lambda m: m.text == "👥 Статистика", state="*")
async def user_stats(message: types.Message):
    if not is_admin(message.from_user.id): return
    process_msg = await show_processing(message, "⏳ Собираю статистику...")
    paid_users = load_paid_users()
    
    active_users = set()
    if os.path.exists(DRAFTS_DIR):
        for filename in os.listdir(DRAFTS_DIR):
            if filename.endswith("_drafts.json"):
                user_id = filename.split('_')[0]
                active_users.add(user_id)
            
    invoice_count = 0
    if os.path.exists(INVOICE_DIR):
        invoice_count = len([f for f in os.listdir(INVOICE_DIR) if f.endswith('.xlsx')])
            
    stats = (
        f"📊 *Статистика бота:*\n\n"
        f"💰 Платные пользователи: *{len(paid_users)}*\n"
        f"👤 Активные пользователи (с черновиками): *{len(active_users)}*\n"
        f"📄 Всего создано накладных: *{invoice_count}*"
    )
    await process_msg.delete()
    await message.answer(stats, parse_mode="Markdown")

@dp.message_handler(lambda m: m.text == "📢 Рассылка", state="*")
async def start_mailing(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await message.answer("Введите текст для рассылки всем пользователям:")
    await AdminStates.mailing_message.set()

@dp.message_handler(state=AdminStates.mailing_message)
async def process_mailing_message(message: types.Message, state: FSMContext):
    users = load_paid_users()
    if not users:
        await message.answer("Список пользователей для рассылки пуст.")
        await state.finish()
        return

    success_count = 0
    fail_count = 0
    process_msg = await show_processing(message, f"Начинаю рассылку для {len(users)} пользователей...")
    
    for user_id in users:
        try:
            await bot.send_message(user_id, message.text)
            success_count += 1
            await asyncio.sleep(0.1)
        except Exception as e:
            fail_count += 1
            logging.warning(f"Не удалось отправить сообщение пользователю {user_id}: {e}")
            
    await process_msg.delete()
    await message.answer(f"✅ Рассылка завершена!\n\nУспешно отправлено: {success_count}\nНе удалось отправить: {fail_count}")
    await state.finish()

@dp.message_handler(lambda m: m.text == "📊 Логи AI", state="*")
async def get_ai_logs(message: types.Message):
    if not is_admin(message.from_user.id): return
    
    process_msg = await show_processing(message, "⏳ Загружаю логи...")
    try:
        log_file = os.path.join(BASE_DIR, "ai_learning_log.json")
        if not os.path.exists(log_file):
            await message.answer("Файл логов AI пока не создан.")
            return
        
        with open(log_file, 'r', encoding='utf-8') as f:
            logs = json.load(f)
        
        if not logs:
            await message.answer("Логи AI пусты.")
            return

        response = "📝 *Последние 5 записей из логов AI:*\n\n"
        for log in logs[-5:]:
            feedback = log.get('feedback', 'нет')
            response += f"📅 *{log['timestamp']}*\n"
            response += f"❓ *Вопрос:* {log['question']}\n"
            response += f"💡 *Ответ:* {log['response'][:100]}...\n"
            response += f"👍 *Фидбек:* {feedback}\n\n"

        await message.answer(response, parse_mode="Markdown")
    finally:
        await process_msg.delete()
    
@dp.message_handler(lambda m: m.text == "🔧 Вкл/Выкл тех. режим", state="*")
async def toggle_maintenance(message: types.Message):
    if not is_admin(message.from_user.id): return
    current_status = is_maintenance_mode_on()
    new_status = not current_status
    set_maintenance_mode(new_status)
    status_text = "🟢 ВКЛЮЧЕН" if new_status else "🔴 ВЫКЛЮЧЕН"
    await message.answer(f"⚙️ Режим технического обслуживания теперь *{status_text}*.", parse_mode="Markdown")
    
@dp.message_handler(lambda m: m.text == "🔄 Перезагрузка", state="*")
async def reboot_bot(message: types.Message):
    if not is_admin(message.from_user.id): return
    await message.answer("🤖 Бот перезапускается... (симуляция)")
    await asyncio.sleep(1)
    await message.answer("Для реального перезапуска бота на сервере используйте команду, предоставленную вашим хостингом (например, `systemctl restart bot.service`).")

@dp.message_handler(lambda m: m.text == "💰 Платежи", state="*")
async def payments_info(message: types.Message):
    if not is_admin(message.from_user.id): return
    await message.answer("Функция управления платежами находится в разработке.\n\nНа данный момент, для добавления платного пользователя, его ID необходимо вручную добавить в файл `paid_users.json`.")

if __name__ == "__main__":
    dp.middleware.setup(MaintenanceMiddleware())
    logging.info("🤖 Бот запускается...")
    executor.start_polling(dp, skip_updates=True)